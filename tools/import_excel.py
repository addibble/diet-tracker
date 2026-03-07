#!/usr/bin/env python3
"""
Excel bulk importer for diet tracker.

Reads an Excel file with a foods sheet and one or more diet log sheets,
identifies recipe rows by formula detection (cells containing =C18*(220/B18)
style expressions), interviews you to confirm or edit ingredient mappings, then
imports everything via the diet tracker API.

Usage:
    python tools/import_excel.py mydata.xlsx
    python tools/import_excel.py mydata.xlsx --api https://diettracker.kndyman.com
    python tools/import_excel.py mydata.xlsx --dry-run   # parse only, no import
    python tools/import_excel.py mydata.xlsx --chol-sodium-unit g  # if stored in grams

Requirements (install into backend venv or any Python 3.10+ env):
    pip install openpyxl requests
"""

import argparse
import datetime
import getpass
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    import requests
except ImportError:
    print("Missing dependencies. Run: pip install openpyxl requests")
    sys.exit(1)


# ── column letter helpers ──────────────────────────────────────────────────────


def col_letter_to_index(letter: str) -> int:
    """A→1, B→2, ..., Z→26, AA→27, ..."""
    result = 0
    for ch in letter.upper().lstrip("$"):
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def col_index_to_letter(idx: int) -> str:
    """1→A, 2→B, ..."""
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(rem + ord("A")) + result
    return result


# ── formula parser ─────────────────────────────────────────────────────────────

# Matches a single ingredient term: [$]COLROW*(GRAMS/[$]COLROW)
# Handles optional $ anchors like $B40, and decimal gram amounts.
_INGREDIENT_RE = re.compile(
    r"\$?([A-Z]+)\$?(\d+)\s*\*\s*\(\s*(\d+(?:\.\d+)?)\s*/\s*\$?([A-Z]+)\$?(\d+)\s*\)",
    re.IGNORECASE,
)


def parse_formula_ingredients(formula: str) -> list[tuple[int, float]]:
    """
    Parse a formula like =C18*(220/B18) + C32*(60/B32) or =C40*(80/$B40) into
    [(18, 220.0), (32, 60.0)].

    Each term is: <macro_col><row> * (<grams> / <serving_col><row>)
    Both cell references must be in the same row. Returns [(row_num, grams), ...].
    """
    results = []
    seen_rows: set[int] = set()
    for m in _INGREDIENT_RE.finditer(formula):
        _macro_col, macro_row, grams, _serving_col, serving_row = m.groups()
        if macro_row == serving_row:
            row_num = int(macro_row)
            if row_num not in seen_rows:
                seen_rows.add(row_num)
                results.append((row_num, float(grams)))
    return results


# ── keyword maps for auto-detecting column headers ────────────────────────────

_MACRO_KEYWORDS: dict[str, str] = {
    "calorie": "calories_per_serving",
    "kcal": "calories_per_serving",
    "fat": "fat_per_serving",
    "saturated": "saturated_fat_per_serving",
    "sat fat": "saturated_fat_per_serving",
    "cholesterol": "cholesterol_per_serving",
    "sodium": "sodium_per_serving",
    "carb": "carbs_per_serving",
    "fiber": "fiber_per_serving",
    "fibre": "fiber_per_serving",
    "protein": "protein_per_serving",
}

_LOG_KEYWORDS: dict[str, str] = {
    "date": "date",
    "meal": "meal_type",
    "type": "meal_type",
    "food": "food_name",
    "item": "food_name",
    "amount": "amount_grams",
    "grams": "amount_grams",
    "weight": "amount_grams",
    "qty": "amount_grams",
    "serving": "amount_grams",
}

# Rows in the log sheets to skip (summary/header sentinel values)
_LOG_SKIP_NAMES = {"food", "item", "name", "food name", "recipe", "totals", "total"}


# ── interactive helpers ────────────────────────────────────────────────────────


def prompt(msg: str, default: str = "") -> str:
    if default:
        val = input(f"  {msg} [{default}]: ").strip()
        return val if val else default
    return input(f"  {msg}: ").strip()


def prompt_yes_no(msg: str, default: bool = True) -> bool:
    yn = "Y/n" if default else "y/N"
    while True:
        s = input(f"  {msg} [{yn}]: ").strip().lower()
        if not s:
            return default
        if s in ("y", "yes"):
            return True
        if s in ("n", "no"):
            return False


def pick_sheet(question: str, sheets: list[str]) -> str:
    if len(sheets) == 1:
        print(f"  (Using only sheet: {sheets[0]})")
        return sheets[0]
    print(f"\n{question}")
    for i, s in enumerate(sheets, 1):
        print(f"  {i}. {s}")
    while True:
        raw = input(f"  Enter number (1-{len(sheets)}): ").strip()
        try:
            idx = int(raw) - 1
            if 0 <= idx < len(sheets):
                return sheets[idx]
        except ValueError:
            pass


def pick_sheets_multi(question: str, sheets: list[str]) -> list[str]:
    """Let the user pick one or more sheets (comma-separated numbers)."""
    print(f"\n{question}")
    for i, s in enumerate(sheets, 1):
        print(f"  {i}. {s}")
    while True:
        raw = input(f"  Enter number(s), comma-separated (e.g. 1,2): ").strip()
        try:
            indices = [int(x.strip()) - 1 for x in raw.split(",")]
            if all(0 <= i < len(sheets) for i in indices) and indices:
                return [sheets[i] for i in indices]
        except ValueError:
            pass
        print("  Invalid selection — try again.")


def show_rows(ws, start_row: int, end_row: int, max_cols: int = 12) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        cells = [str(v)[:14] if v is not None else "" for v in row[:max_cols]]
        if any(cells):
            print("  | " + " | ".join(f"{c:<14}" for c in cells) + " |")


# ── API client ────────────────────────────────────────────────────────────────


class DietTrackerAPI:
    def __init__(self, base_url: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()

    def login(self, password: str) -> bool:
        r = self.session.post(f"{self.base_url}/api/auth/login", json={"password": password})
        return r.status_code == 200

    def list_foods(self) -> list[dict]:
        r = self.session.get(f"{self.base_url}/api/foods")
        r.raise_for_status()
        return r.json()

    def create_food(self, data: dict) -> dict:
        r = self.session.post(f"{self.base_url}/api/foods", json=data)
        r.raise_for_status()
        return r.json()

    def list_recipes(self) -> list[dict]:
        r = self.session.get(f"{self.base_url}/api/recipes")
        r.raise_for_status()
        return r.json()

    def create_recipe(self, data: dict) -> dict:
        r = self.session.post(f"{self.base_url}/api/recipes", json=data)
        r.raise_for_status()
        return r.json()

    def create_meal(self, data: dict) -> dict:
        r = self.session.post(f"{self.base_url}/api/meals", json=data)
        r.raise_for_status()
        return r.json()


# ── main importer ─────────────────────────────────────────────────────────────


class Importer:
    def __init__(
        self,
        excel_path: str,
        api: DietTrackerAPI,
        dry_run: bool = False,
        chol_sodium_unit: str = "mg",  # "mg" or "g"
    ) -> None:
        self.api = api
        self.dry_run = dry_run
        # Multiplier to convert spreadsheet chol/sodium values to mg
        self.chol_sod_factor = 1000.0 if chol_sodium_unit == "g" else 1.0

        print(f"\nOpening {excel_path} ...")
        self.wb_f = openpyxl.load_workbook(excel_path, data_only=False)
        self.wb_v = openpyxl.load_workbook(excel_path, data_only=True)

        # row_num → food dict
        self.food_rows: dict[int, dict] = {}
        # row_num → API food_id (positive) or -recipe_id (negative)
        self.row_to_id: dict[int, int] = {}

    # ── top-level flow ─────────────────────────────────────────────────────────

    def run(self) -> None:
        sheets = self.wb_f.sheetnames
        print(f"Sheets: {', '.join(sheets)}")

        foods_name = pick_sheet("Which sheet has your FOODS list?", sheets)
        log_names = pick_sheets_multi("Which sheet(s) have your DIET LOG? (select all that apply)", sheets)

        foods_ws_f = self.wb_f[foods_name]
        foods_ws_v = self.wb_v[foods_name]

        print("\n" + "=" * 60)
        print("STEP 1: Foods — column mapping")
        print("=" * 60)
        foods_col_map = self._map_foods_columns(foods_ws_f)
        self._parse_food_rows(foods_ws_f, foods_ws_v, foods_col_map)

        print("\n" + "=" * 60)
        print("STEP 2: Recipes — confirm formula-detected ingredients")
        print("=" * 60)
        confirmed_recipes = self._interview_recipes(foods_ws_f, foods_col_map)

        if not self.dry_run:
            print("\n" + "=" * 60)
            print("STEP 3: Importing foods and recipes")
            print("=" * 60)
            self._import_foods(confirmed_recipes)

        print("\n" + "=" * 60)
        print("STEP 4: Diet log — column mapping")
        print("=" * 60)
        # Use the first log sheet to map columns (assume same layout for all)
        first_log_ws = self.wb_v[log_names[0]]
        log_col_map = self._map_log_columns(first_log_ws)

        all_entries: list[dict] = []
        for log_name in log_names:
            log_ws = self.wb_v[log_name]
            print(f"\n  Parsing '{log_name}'...")
            entries = self._parse_log(log_ws, log_col_map, source=log_name)
            all_entries.extend(entries)

        print(f"\n  Total log entries across all sheets: {len(all_entries)}")

        if not self.dry_run:
            print("\n" + "=" * 60)
            print("STEP 5: Importing meal log")
            print("=" * 60)
            self._import_log(all_entries)

        print("\n✓ Done!")

    # ── foods column mapping ───────────────────────────────────────────────────

    def _map_foods_columns(self, ws) -> dict[str, int]:
        print("\nFirst rows of the foods sheet:")
        show_rows(ws, 1, 4)

        col_map: dict[str, int] = {}
        for row_idx in range(1, 4):
            row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), None)
            if not row or len([v for v in row if v is not None]) < 2:
                continue
            for col_idx, header in enumerate(row, 1):
                if header is None:
                    continue
                h = str(header).lower().strip()
                if h in ("name", "food", "item", "food name") and "name" not in col_map:
                    col_map["name"] = col_idx
                elif h in ("serving", "serving size", "grams/serving", "size (g)", "weight (g)") and "serving_size_grams" not in col_map:
                    col_map["serving_size_grams"] = col_idx
                else:
                    for kw, field in _MACRO_KEYWORDS.items():
                        if kw in h and field not in col_map:
                            col_map[field] = col_idx
                            break
            break

        print("\nConfirm or enter column letters (e.g. A, B, C). Press Enter to keep detected value.\n")

        required = [
            ("name", "Food name"),
            ("serving_size_grams", "Serving size (grams)"),
            ("calories_per_serving", "Calories"),
            ("fat_per_serving", "Total fat (g)"),
            ("carbs_per_serving", "Carbohydrates (g)"),
            ("protein_per_serving", "Protein (g)"),
        ]
        optional = [
            ("saturated_fat_per_serving", "Saturated fat (g) [optional]"),
            ("cholesterol_per_serving", "Cholesterol [optional]"),
            ("sodium_per_serving", "Sodium [optional]"),
            ("fiber_per_serving", "Fiber (g) [optional]"),
        ]

        for field, label in required + optional:
            default = col_index_to_letter(col_map[field]) if field in col_map else ""
            val = prompt(label, default)
            if val.strip():
                col_map[field] = col_letter_to_index(val.strip())

        return col_map

    # ── parse all food rows ────────────────────────────────────────────────────

    def _parse_food_rows(self, ws_f, ws_v, col_map: dict[str, int]) -> None:
        name_col = col_map.get("name")
        cal_col = col_map.get("calories_per_serving")
        if not name_col:
            print("ERROR: name column not configured.")
            return

        print("\nParsing food rows...")
        recipe_count = 0

        for row_idx in range(1, ws_f.max_row + 1):
            name_v = ws_v.cell(row=row_idx, column=name_col).value
            if not name_v or not str(name_v).strip():
                continue
            name = str(name_v).strip()
            if name.lower() in ("name", "food", "item", "food name"):
                continue

            # Detect recipe: calories cell is a formula
            formula_str = ""
            is_recipe = False
            if cal_col:
                raw = ws_f.cell(row=row_idx, column=cal_col).value
                if isinstance(raw, str) and raw.startswith("="):
                    is_recipe = True
                    formula_str = raw

            food: dict = {"name": name, "is_recipe": is_recipe, "formula": formula_str, "row": row_idx}

            for field, col_idx in col_map.items():
                if field == "name":
                    continue
                val = ws_v.cell(row=row_idx, column=col_idx).value
                try:
                    fval = float(val) if val is not None else 0.0
                except (TypeError, ValueError):
                    fval = 0.0
                # Convert cholesterol and sodium from grams to mg if needed
                if field in ("cholesterol_per_serving", "sodium_per_serving"):
                    fval *= self.chol_sod_factor
                food[field] = fval

            self.food_rows[row_idx] = food
            if is_recipe:
                recipe_count += 1

        plain = len(self.food_rows) - recipe_count
        print(f"  Found {len(self.food_rows)} rows: {plain} plain foods, {recipe_count} potential recipes.")

    # ── recipe interview ───────────────────────────────────────────────────────

    def _interview_recipes(self, ws_f, col_map: dict[str, int]) -> dict[int, list[tuple[int, float]]]:
        recipe_rows = [r for r, f in self.food_rows.items() if f["is_recipe"]]
        confirmed: dict[int, list[tuple[int, float]]] = {}

        if not recipe_rows:
            print("  No formula-based rows detected — nothing to review.")
            return confirmed

        cal_col = col_map.get("calories_per_serving")
        plain_foods = {r: f for r, f in self.food_rows.items() if not f["is_recipe"]}

        print(f"\n  {len(recipe_rows)} rows have formula-based macros (potential recipes).")
        print("  I'll show each one so you can confirm or correct the ingredients.\n")

        for row_idx in recipe_rows:
            food = self.food_rows[row_idx]
            print(f"{'─'*60}")
            print(f"  Row {row_idx}: \"{food['name']}\"")

            formula = (ws_f.cell(row=row_idx, column=cal_col).value or "") if cal_col else food.get("formula", "")
            print(f"  Calories formula: {formula}")

            parsed = parse_formula_ingredients(str(formula)) if formula else []

            if parsed:
                print(f"\n  Auto-detected {len(parsed)} ingredient(s):")
                for ing_row, grams in parsed:
                    ing = self.food_rows.get(ing_row)
                    ing_name = ing["name"] if ing else f"(row {ing_row} — not in foods sheet)"
                    print(f"    Row {ing_row:3d}  {grams:6.0f} g   {ing_name}")
            else:
                print("  (Could not parse ingredients from formula automatically.)")

            is_recipe = prompt_yes_no(f"\n  Treat this as a RECIPE?")
            if not is_recipe:
                food["is_recipe"] = False
                continue

            if parsed and prompt_yes_no("  Are the detected ingredients correct?"):
                confirmed[row_idx] = parsed
                continue

            # Manual ingredient entry
            print("\n  Enter ingredients manually.")
            print("  Plain foods available:")
            for r, f in sorted(plain_foods.items()):
                print(f"    Row {r:3d}: {f['name']}")

            manual: list[tuple[int, float]] = []
            while True:
                row_s = input("\n  Add ingredient — row number (or Enter to finish): ").strip()
                if not row_s:
                    break
                try:
                    ing_row = int(row_s)
                except ValueError:
                    print("    Not a number.")
                    continue
                if ing_row not in self.food_rows:
                    print(f"    Row {ing_row} not in foods list.")
                    continue
                grams_s = input(f"  Grams of {self.food_rows[ing_row]['name']}: ").strip()
                try:
                    manual.append((ing_row, float(grams_s)))
                except ValueError:
                    print("    Invalid grams value.")

            confirmed[row_idx] = manual

        print(f"\n  Confirmed {len(confirmed)} recipe(s).")
        return confirmed

    # ── import foods + recipes ─────────────────────────────────────────────────

    def _import_foods(self, confirmed_recipes: dict[int, list[tuple[int, float]]]) -> None:
        recipe_row_set = set(confirmed_recipes.keys())

        try:
            existing_foods = {f["name"].lower(): f["id"] for f in self.api.list_foods()}
            existing_recipes = {r["name"].lower(): r["id"] for r in self.api.list_recipes()}
            print(f"  Database has {len(existing_foods)} food(s) and {len(existing_recipes)} recipe(s) already.")
        except Exception as e:
            print(f"  Warning: could not fetch existing items — {e}")
            existing_foods = {}
            existing_recipes = {}

        # ── plain foods ────────────────────────────────────────────────────────
        plain_rows = [r for r in self.food_rows if r not in recipe_row_set]
        print(f"\n  Importing {len(plain_rows)} plain food(s)...")
        imported = skipped = errors = 0

        for row_idx in plain_rows:
            food = self.food_rows[row_idx]
            name_lower = food["name"].lower()
            if name_lower in existing_foods:
                self.row_to_id[row_idx] = existing_foods[name_lower]
                skipped += 1
                continue

            payload = {
                "name": food["name"],
                "serving_size_grams": food.get("serving_size_grams") or 100.0,
                "calories_per_serving": food.get("calories_per_serving") or 0.0,
                "fat_per_serving": food.get("fat_per_serving") or 0.0,
                "saturated_fat_per_serving": food.get("saturated_fat_per_serving") or 0.0,
                "cholesterol_per_serving": food.get("cholesterol_per_serving") or 0.0,
                "sodium_per_serving": food.get("sodium_per_serving") or 0.0,
                "carbs_per_serving": food.get("carbs_per_serving") or 0.0,
                "fiber_per_serving": food.get("fiber_per_serving") or 0.0,
                "protein_per_serving": food.get("protein_per_serving") or 0.0,
            }
            try:
                result = self.api.create_food(payload)
                self.row_to_id[row_idx] = result["id"]
                imported += 1
                print(f"    ✓ {food['name']} → food id={result['id']}")
            except Exception as e:
                errors += 1
                print(f"    ✗ {food['name']}: {e}")

        print(f"  Foods: {imported} imported, {skipped} already existed, {errors} error(s).")

        # ── recipes ────────────────────────────────────────────────────────────
        print(f"\n  Importing {len(confirmed_recipes)} recipe(s)...")
        r_imported = r_skipped = r_errors = 0

        for row_idx, ingredients in confirmed_recipes.items():
            recipe_name = self.food_rows[row_idx]["name"]
            name_lower = recipe_name.lower()
            if name_lower in existing_recipes:
                self.row_to_id[row_idx] = -existing_recipes[name_lower]
                r_skipped += 1
                continue

            components = []
            for ing_row, grams in ingredients:
                food_id = self.row_to_id.get(ing_row)
                if food_id is None:
                    print(f"    ⚠ Ingredient row {ing_row} ({self.food_rows.get(ing_row, {}).get('name', '?')}) not imported — skipping.")
                    continue
                if food_id < 0:
                    print(f"    ⚠ Ingredient row {ing_row} is itself a recipe (nesting not supported) — skipping.")
                    continue
                components.append({"food_id": food_id, "amount_grams": grams})

            if not components:
                print(f"    ✗ {recipe_name}: no valid components, skipping.")
                r_errors += 1
                continue

            try:
                result = self.api.create_recipe({"name": recipe_name, "components": components})
                self.row_to_id[row_idx] = -result["id"]
                r_imported += 1
                print(f"    ✓ {recipe_name} → recipe id={result['id']}")
            except Exception as e:
                r_errors += 1
                print(f"    ✗ {recipe_name}: {e}")

        print(f"  Recipes: {r_imported} imported, {r_skipped} already existed, {r_errors} error(s).")

    # ── log column mapping ─────────────────────────────────────────────────────

    def _map_log_columns(self, ws) -> dict[str, int]:
        print("\nFirst rows of the diet log sheet:")
        show_rows(ws, 1, 4)

        col_map: dict[str, int] = {}
        for row_idx in range(1, 4):
            row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), None)
            if not row or len([v for v in row if v is not None]) < 2:
                continue
            for col_idx, header in enumerate(row, 1):
                if header is None:
                    continue
                h = str(header).lower().strip()
                for kw, field in _LOG_KEYWORDS.items():
                    if kw in h and field not in col_map:
                        col_map[field] = col_idx
                        break
            break

        print("\nConfirm or enter column letters.\n")
        fields = [
            ("date", "Date column"),
            ("meal_type", "Meal type column [optional — leave blank to default to 'meal']"),
            ("food_name", "Food/recipe name column"),
            ("amount_grams", "Amount in grams column"),
        ]
        for field, label in fields:
            default = col_index_to_letter(col_map[field]) if field in col_map else ""
            val = prompt(label, default)
            if val.strip():
                col_map[field] = col_letter_to_index(val.strip())

        return col_map

    # ── parse log entries ──────────────────────────────────────────────────────

    def _parse_log(self, ws, col_map: dict[str, int], source: str = "") -> list[dict]:
        date_col = col_map.get("date")
        meal_type_col = col_map.get("meal_type")
        food_name_col = col_map.get("food_name")
        amount_col = col_map.get("amount_grams")

        if not food_name_col:
            print("  ERROR: food name column not configured.")
            return []

        food_by_name = {f["name"].lower(): r for r, f in self.food_rows.items()}
        unmatched: set[str] = set()
        entries: list[dict] = []

        for row in ws.iter_rows(values_only=True):
            food_name_raw = row[food_name_col - 1]
            if not food_name_raw:
                continue
            food_name = str(food_name_raw).strip()
            if food_name.lower() in _LOG_SKIP_NAMES:
                continue

            date_raw = row[date_col - 1] if date_col else None
            log_date = _parse_date(date_raw)
            if log_date is None:
                continue

            meal_type_raw = row[meal_type_col - 1] if meal_type_col else None
            meal_type = _normalize_meal_type(meal_type_raw)

            amount_raw = row[amount_col - 1] if amount_col else None
            try:
                amount_grams = float(amount_raw) if amount_raw is not None else 100.0
            except (TypeError, ValueError):
                amount_grams = 100.0

            # Case-insensitive food name lookup
            food_row = food_by_name.get(food_name.lower())
            if food_row is None:
                unmatched.add(food_name)

            entries.append({
                "date": log_date,
                "meal_type": meal_type,
                "food_name": food_name,
                "food_row": food_row,
                "amount_grams": amount_grams,
                "source": source,
            })

        print(f"    {len(entries)} entries parsed.")
        if unmatched:
            print(f"    ⚠ {len(unmatched)} food name(s) not matched to foods sheet:")
            for name in sorted(unmatched):
                print(f"      - {name}")

        return entries

    # ── import log ─────────────────────────────────────────────────────────────

    def _import_log(self, entries: list[dict]) -> None:
        # Group by (date, meal_type) — items on the same date+type become one MealLog
        groups: dict[tuple, list[dict]] = defaultdict(list)
        for entry in entries:
            groups[(entry["date"], entry["meal_type"])].append(entry)

        print(f"  Importing {len(groups)} meal log(s) ({len(entries)} items)...")
        imported = errors = skipped_items = 0

        for (log_date, meal_type), items in sorted(groups.items()):
            meal_items = []
            for item in items:
                food_row = item["food_row"]
                if food_row is None:
                    skipped_items += 1
                    continue
                id_val = self.row_to_id.get(food_row)
                if id_val is None:
                    skipped_items += 1
                    continue
                if id_val < 0:
                    meal_items.append({"recipe_id": -id_val, "amount_grams": item["amount_grams"]})
                else:
                    meal_items.append({"food_id": id_val, "amount_grams": item["amount_grams"]})

            if not meal_items:
                continue

            try:
                self.api.create_meal({
                    "date": log_date.isoformat(),
                    "meal_type": meal_type,
                    "items": meal_items,
                })
                imported += 1
            except Exception as e:
                errors += 1
                print(f"    ✗ {log_date} {meal_type}: {e}")

        print(f"  Meal logs: {imported} imported, {errors} error(s).")
        if skipped_items:
            print(f"  {skipped_items} item(s) skipped (unmatched food names).")


# ── date / meal-type helpers ──────────────────────────────────────────────────


def _parse_date(raw) -> datetime.date | None:
    if isinstance(raw, datetime.datetime):
        return raw.date()
    if isinstance(raw, datetime.date):
        return raw
    if raw is None:
        return None
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _normalize_meal_type(raw) -> str:
    if raw is None:
        return "meal"
    s = str(raw).lower().strip()
    if "break" in s:
        return "breakfast"
    if "lunch" in s:
        return "lunch"
    if "dinner" in s or "supper" in s:
        return "dinner"
    if "snack" in s:
        return "snack"
    return "meal"


# ── entry point ───────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import historical diet data from an Excel file into Diet Tracker."
    )
    parser.add_argument("excel_file", help="Path to .xlsx file")
    parser.add_argument(
        "--api",
        default="http://localhost:8000",
        help="Diet Tracker API base URL (default: http://localhost:8000)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and review without making any API calls",
    )
    parser.add_argument(
        "--chol-sodium-unit",
        choices=["mg", "g"],
        default="mg",
        help="Unit for cholesterol and sodium in the spreadsheet: 'mg' (default) or 'g' (auto-×1000)",
    )
    args = parser.parse_args()

    if not Path(args.excel_file).exists():
        print(f"ERROR: File not found: {args.excel_file}")
        sys.exit(1)

    api = DietTrackerAPI(args.api)

    if not args.dry_run:
        print(f"\nConnecting to {args.api} ...")
        password = getpass.getpass("Diet Tracker password: ")
        if not api.login(password):
            print("ERROR: Login failed. Check password and API URL.")
            sys.exit(1)
        print("Logged in successfully.")
    else:
        print("\n[DRY RUN — no data will be imported]")

    importer = Importer(
        args.excel_file,
        api,
        dry_run=args.dry_run,
        chol_sodium_unit=args.chol_sodium_unit,
    )
    importer.run()


if __name__ == "__main__":
    main()
